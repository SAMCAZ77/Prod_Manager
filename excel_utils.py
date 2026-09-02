"""
excel_utils.py
Handles exporting table data to nicely formatted .xlsx files using openpyxl.
"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1E3C6E", end_color="1E3C6E", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
TITLE_FONT = Font(bold=True, size=14, color="1E3C6E")


def export_to_excel(path, title, headers, rows, sheet_name="Data"):
    """
    Write `rows` (list of tuples/lists) with `headers` to an .xlsx file at
    `path`, with a title row, styled header row, borders and auto-sized
    columns. Returns the path on success.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] if sheet_name else "Data"

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left")

    subtitle_cell = ws.cell(
        row=2, column=1,
        value=f"Exported on {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    )
    subtitle_cell.font = Font(italic=True, size=9, color="808080")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    header_row_idx = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for r_offset, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=header_row_idx + r_offset, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-size columns based on content length
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            if col_idx - 1 < len(row) and row[col_idx - 1] is not None:
                max_len = max(max_len, len(str(row[col_idx - 1])))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)

    wb.save(path)
    return path
